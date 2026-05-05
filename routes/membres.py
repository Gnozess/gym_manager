from flask import Blueprint, render_template, redirect, url_for, flash, request, Response
from flask_login import login_required, current_user
from extensions import db
from models.membre import Membre
from models.abonnement import Abonnement
from datetime import date
import io

membres_bp = Blueprint('membres', __name__)


def valider_telephone(telephone):
    if not telephone:
        return True, None
    telephone = telephone.strip().replace(' ', '')
    if not telephone.startswith('+226'):
        return False, 'Le numero doit commencer par +226.'
    if len(telephone) != 12:
        return False, 'Le numero doit contenir 12 caracteres (+226XXXXXXXX).'
    if not telephone[4:].isdigit():
        return False, 'Le numero doit contenir uniquement des chiffres apres +226.'
    return True, None


@membres_bp.route('/')
def index():
    return redirect(url_for('membres.dashboard'))


@membres_bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.est_membre:
        return redirect(url_for('abonnements.liste'))

    from sqlalchemy import extract, func
    annee = request.args.get('annee', date.today().year, type=int)

    total_membres = Membre.query.filter_by(actif=True).count()
    abonnements_actifs = Abonnement.query.filter(
        Abonnement.statut == 'actif',
        Abonnement.date_fin >= date.today()
    ).count()
    expirations = Abonnement.query.filter(
        Abonnement.statut == 'actif',
        Abonnement.date_fin >= date.today(),
        Abonnement.date_fin <= date.fromordinal(date.today().toordinal() + 7)
    ).all()

    # Données histogramme — abonnés par mois pour l'année choisie
    resultats = db.session.query(
        extract('month', Abonnement.date_debut).label('mois'),
        func.count(Abonnement.id).label('total')
    ).filter(
        extract('year', Abonnement.date_debut) == annee,
        Abonnement.type_abonnement == 'forfait'
    ).group_by('mois').order_by('mois').all()

    # Initialiser 12 mois à 0
    donnees_mois = [0] * 12
    for r in resultats:
        donnees_mois[int(r.mois) - 1] = r.total

    # Années disponibles pour le filtre
    annees = db.session.query(
        extract('year', Abonnement.date_debut).label('annee')
    ).distinct().order_by('annee').all()
    annees = [int(a.annee) for a in annees]
    if not annees:
        annees = [date.today().year]

    return render_template('dashboard.html',
        total_membres=total_membres,
        abonnements_actifs=abonnements_actifs,
        expirations=expirations,
        donnees_mois=donnees_mois,
        annee=annee,
        annees=annees
    )


@membres_bp.route('/membres')
@login_required
def liste():
    page = request.args.get('page', 1, type=int)
    recherche = request.args.get('q', '')
    filtre_statut = request.args.get('statut', '')
    par_page = 10

    query = Membre.query

    if recherche:
        query = query.filter(
            (Membre.nom.ilike(f'%{recherche}%')) |
            (Membre.prenom.ilike(f'%{recherche}%')) |
            (Membre.email.ilike(f'%{recherche}%')) |
            (Membre.telephone.ilike(f'%{recherche}%'))
        )

    if filtre_statut == 'actif':
        query = query.filter_by(actif=True)
    elif filtre_statut == 'inactif':
        query = query.filter_by(actif=False)

    membres_pagines = query.order_by(Membre.nom.asc()).paginate(
        page=page, per_page=par_page, error_out=False
    )

    return render_template('membres/liste.html',
        membres=membres_pagines,
        recherche=recherche,
        filtre_statut=filtre_statut
    )


@membres_bp.route('/membres/export')
@login_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Module openpyxl manquant. Installez-le avec : pip install openpyxl', 'danger')
        return redirect(url_for('membres.liste'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Membres'

    headers = ['#', 'Nom', 'Prenom', 'Email', 'Telephone',
               'Date inscription', 'Statut', 'Abonnement']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='212529')
        cell.alignment = Alignment(horizontal='center')

    membres = Membre.query.order_by(Membre.nom.asc()).all()
    for row, membre in enumerate(membres, 2):
        ab = membre.abonnement_actif
        ws.cell(row=row, column=1, value=membre.id)
        ws.cell(row=row, column=2, value=membre.nom)
        ws.cell(row=row, column=3, value=membre.prenom)
        ws.cell(row=row, column=4, value=membre.email)
        ws.cell(row=row, column=5, value=membre.telephone or '')
        ws.cell(row=row, column=6,
                value=membre.date_inscription.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=7,
                value='Actif' if membre.actif else 'Inactif')
        if ab:
            ws.cell(row=row, column=8,
                    value=ab.forfait.nom if ab.forfait else 'Journalier')
        else:
            ws.cell(row=row, column=8, value='Aucun')

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=membres.xlsx'}
    )


@membres_bp.route('/membres/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter():
    if request.method == 'POST':
        nom = request.form.get('nom', '').strip().upper()
        prenom = request.form.get('prenom', '').strip().upper()
        email = request.form.get('email', '').strip()
        telephone = request.form.get('telephone', '').strip() or None

        if Membre.query.filter_by(email=email).first():
            flash('Un membre avec cet email existe deja.', 'danger')
            return redirect(url_for('membres.liste'))

        if telephone:
            valide, erreur = valider_telephone(telephone)
            if not valide:
                flash(erreur, 'danger')
                return redirect(url_for('membres.liste'))
            if Membre.query.filter_by(telephone=telephone).first():
                flash('Un membre avec ce numero de telephone existe deja.', 'danger')
                return redirect(url_for('membres.liste'))

        membre = Membre(
            nom=nom,
            prenom=prenom,
            email=email,
            telephone=telephone
        )
        db.session.add(membre)
        db.session.commit()
        flash(f'Membre {prenom} {nom} ajoute avec succes !', 'success')
        return redirect(url_for('membres.liste'))

    return render_template('membres/liste.html')


@membres_bp.route('/membres/modifier/<int:id>', methods=['GET', 'POST'])
@login_required
def modifier(id):
    membre = Membre.query.get_or_404(id)
    if request.method == 'POST':
        nom = request.form.get('nom', '').strip().upper()
        prenom = request.form.get('prenom', '').strip().upper()
        email = request.form.get('email', '').strip()
        telephone = request.form.get('telephone', '').strip() or None
        actif = request.form.get('actif') == 'on'

        if not actif and membre.actif:
            abonnement_actif = Abonnement.query.filter_by(
                membre_id=membre.id,
                type_abonnement='forfait',
                statut='actif'
            ).filter(Abonnement.date_fin >= date.today()).first()

            if abonnement_actif:
                flash(
                    'Impossible de desactiver ce membre : '
                    'il a un abonnement forfait actif.',
                    'danger'
                )
                return redirect(url_for('membres.liste'))

        existing_email = Membre.query.filter_by(email=email).first()
        if existing_email and existing_email.id != membre.id:
            flash('Un membre avec cet email existe deja.', 'danger')
            return redirect(url_for('membres.liste'))

        if telephone:
            valide, erreur = valider_telephone(telephone)
            if not valide:
                flash(erreur, 'danger')
                return redirect(url_for('membres.liste'))
            existing_tel = Membre.query.filter_by(telephone=telephone).first()
            if existing_tel and existing_tel.id != membre.id:
                flash('Un membre avec ce numero de telephone existe deja.', 'danger')
                return redirect(url_for('membres.liste'))

        membre.nom = nom
        membre.prenom = prenom
        membre.email = email
        membre.telephone = telephone
        membre.actif = actif
        db.session.commit()
        flash('Membre modifie avec succes !', 'success')
        return redirect(url_for('membres.liste'))

    return render_template('membres/modifier.html', membre=membre)


@membres_bp.route('/membres/supprimer/<int:id>', methods=['POST'])
@login_required
def supprimer(id):
    membre = Membre.query.get_or_404(id)
    db.session.delete(membre)
    db.session.commit()
    flash(f'Membre {membre.prenom} {membre.nom} supprime.', 'warning')
    return redirect(url_for('membres.liste'))